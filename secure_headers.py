import sys
import re
import warnings

# Suppress SSL warnings
warnings.filterwarnings('ignore', message='Unverified HTTPS request')

# Configuration
USER_AGENT = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
TIMEOUT_SECONDS = 15
HSTS_MIN_AGE = 10368000 # 120 days

# Security Check Configuration
SEC_HEADERS = {
'Strict-Transport-Security': 'error',
'Content-Security-Policy': 'warning',
'X-Frame-Options': 'warning',
'X-Content-Type-Options': 'warning',
'Referrer-Policy': 'warning',
'Permissions-Policy': 'warning'
}


def get_urls_to_try(target):
    """
    Returns a list of URLs to try for a given target.
    If target already has a scheme, use it as-is.
    Otherwise, try HTTPS first, then fall back to HTTP.
    """
    if target.startswith(('http://', 'https://')):
        return [target]
    return ['https://' + target, 'http://' + target]

def check_target(target):
    """
    Attempts to connect to the target using requests library.
    Tries HTTPS first, then falls back to HTTP.
    Returns (final_url, headers_dict) if successful, (None, None) otherwise.
    """
    try:
        import requests
    except ImportError:
        print("Error: requests library not installed. Run: uv pip install requests")
        sys.exit(1)
    
    headers = {'User-Agent': USER_AGENT}
    urls_to_try = get_urls_to_try(target)
    
    for url in urls_to_try:
        try:
            response = requests.get(
                url,
                headers=headers,
                timeout=TIMEOUT_SECONDS,
                verify=False,  # Disable SSL verification
                allow_redirects=True
            )
            # Return the final URL and headers (case-insensitive)
            return response.url, {k.lower(): v for k, v in response.headers.items()}
        except Exception:
            continue  # Try next URL (HTTP fallback)
    
    return None, None  # All attempts failed

def validate_hsts(value):
    """
    Validates HSTS header.
    Requirement: max-age >= 120 days.
    Note: includeSubDomains is checked but not enforced as failure (per user request).
    """
    match = re.search(r'max-age=(\d+)', value.lower())
    if match:
        max_age = int(match.group(1))
        if max_age < HSTS_MIN_AGE:
            return True, f"max-age too low ({max_age})"
        return False, None
    return True, "invalid/missing max-age"

def validate_csp(value):
    """
    Validates CSP header.
    Rules:
    1. unsafe-eval is always misconfigured in script/default src (wasm-unsafe-eval is ignored).
    2. unsafe-inline is misconfigured in script/default src UNLESS nonce or hash is present.
    """
    csp_issues = []
    directives = value.split(';')
    parsed_directives = {}
    for directive in directives:
        directive = directive.strip()
        parts = directive.split()
        if not parts: continue
        d_name = parts[0].lower()
        parsed_directives[d_name] = parts[1:]
    
    # Determine effective script-src
    target_values = []
    if 'script-src' in parsed_directives:
        target_values = parsed_directives['script-src']
    elif 'default-src' in parsed_directives:
        target_values = parsed_directives['default-src']
    
    if target_values:
        # Check for unsafe-inline
        if "'unsafe-inline'" in target_values:
            has_nonce_or_hash = any(v.startswith(("'nonce-", "'sha")) for v in target_values)
            if not has_nonce_or_hash and "unsafe-inline" not in csp_issues:
                csp_issues.append("unsafe-inline")
        # Check for unsafe-eval (strictly bad, distinct from wasm-unsafe-eval)
        if "'unsafe-eval'" in target_values and "unsafe-eval" not in csp_issues:
            csp_issues.append("unsafe-eval")
            
    if csp_issues:
        return True, ", ".join(csp_issues)
    return False, None

def validate_header(header_name, value):
    """
    Dispatcher for header specific validation logic.
    Returns (is_misconfigured: bool, misconf_value: str | None)
    """
    header_lower = header_name.lower()
    if header_lower == 'x-xss-protection' and value == '0':
        return True, "0"
    elif header_lower == 'referrer-policy':
        if 'unsafe-url' in value.lower():
            return True, "unsafe-url"
    elif header_lower == 'strict-transport-security':
        return validate_hsts(value)
    elif header_lower == 'content-security-policy':
        return validate_csp(value)
    elif header_lower == 'x-frame-options':
        # Strict enforcement: Must be DENY or SAMEORIGIN
        if value.upper() not in ['DENY', 'SAMEORIGIN']:
            return True, f"bad value: {value}"
    elif header_lower == 'x-content-type-options':
        if 'nosniff' not in value.lower():
            return True, f"bad value: {value}"
    return False, None


def save_to_excel(results, filename):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    except ImportError:
        print("Error: openpyxl not installed.")
        return
        
    wb = Workbook()
    ws = wb.active
    ws.title = "Security Headers Report"
    
    # Styles
    fill_ok = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    fill_missing = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    fill_misconfigured = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    header_font = Font(bold=True)
    center_aligned = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # write Header Row
    headers_list = list(SEC_HEADERS.keys())
    ws.append(["Input Domain", "Final URL"] + headers_list)
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = center_aligned
        cell.border = thin_border
        
    row_idx = ws.max_row + 1
    for input_domain, data in results.items():
        final_url = data.get("final_url", input_domain)
        ws.cell(row=row_idx, column=1, value=input_domain).border = thin_border
        ws.cell(row=row_idx, column=2, value=final_url).border = thin_border
        present = data.get("present", {})
        missing = data.get("missing", [])
        
        for col_idx, header_name in enumerate(headers_list, start=3):
            cell_value = "N/A"
            cell_fill = None
            
            # Check if this host failed to connect
            if data.get("failed"):
                cell_value = "Connection Failed"
                # Leave cell_fill as None (no color) for failed connections
            elif header_name in missing:
                cell_value = "Missing"
                cell_fill = fill_missing
            elif header_name in present:
                value = present[header_name]
                is_bad, msg = validate_header(header_name, value)
                if is_bad:
                    cell_value = msg if msg else value
                    cell_fill = fill_misconfigured
                else:
                    cell_value = "OK"
                    cell_fill = fill_ok
            cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
            cell.alignment = center_aligned
            cell.border = thin_border
            if cell_fill:
                cell.fill = cell_fill
        row_idx += 1
        
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                val_len = len(str(cell.value))
                if val_len > max_length:
                    max_length = min(val_len, 50)
            except:
                pass
        ws.column_dimensions[column_letter].width = max_length + 2
        
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    wb.save(filename)

def main():
    if len(sys.argv) != 3:
        print("Usage: python secure_headers.py <hosts_file> <output_xlsx>")
        sys.exit(1)
        
    hosts_file = sys.argv[1]
    output_file = sys.argv[2]
    
    try:
        with open(hosts_file, 'r') as f:
            targets = f.read().splitlines()
    except FileNotFoundError:
        print(f"Error: File {hosts_file} not found.")
        sys.exit(1)
        
    json_out = {}
    valid_targets = [t for t in targets if t.strip()]
    total_targets = len(valid_targets)
    
    # Progress bar setup
    pbar = None
    try:
        from tqdm import tqdm
        pbar = tqdm(total=total_targets, unit="site")
    except ImportError:
        print("Processing targets...")
        
    processed_count = 0
    for target in valid_targets:
        rUrl, headers = check_target(target)
        
        # Update progress
        if pbar:
            pbar.update(1)
            pbar.set_description(f"Scanning {target[:20]}")
        else:
            processed_count += 1
            print(f"[{processed_count}/{total_targets}] Scanning {target}...", end='\r')
            
        if rUrl is None:
            # Include failed hosts in output with connection failed status
            json_out[target] = {"final_url": target, "present": {}, "missing": [], "failed": True}
            continue
        
        json_results = {"final_url": rUrl, "present": {}, "missing": []}
            
        for header_name in SEC_HEADERS:
            header_lower = header_name.lower()
            if header_lower in headers:
                json_results["present"][header_name] = headers[header_lower]
            else:
                # HSTS only required on HTTPS
                if header_name == 'Strict-Transport-Security' and not rUrl.startswith('https://'):
                    continue
                json_results["missing"].append(header_name)
        
        json_out[target] = json_results
        
    if pbar:
        pbar.close()
    else:
        print("\nScan complete.")
        
    save_to_excel(json_out, output_file)

if __name__ == "__main__":
    main()